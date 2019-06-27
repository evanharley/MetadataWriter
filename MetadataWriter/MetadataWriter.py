import os, sys
import openpyxl
import exiftool
import pickle
from tkinter import filedialog
from pprint import pprint

class MetadataWriter():
    
    def __init__(self, *args, **kwargs):
        self.fn = filedialog.askopenfilename(title='Open', 
                                             defaultextension='.xlsx', filetypes=[('Excel Files', '*.xlsx')])
        self.head, self.tail = os.path.split(self.fn)
        self.directory = self.head
        try:
            self.spreadsheet_data = pickle.load(open(self.tail[:self.tail.find('.')] + '.pkl', 'rb'))
        except FileNotFoundError:
            f = openpyxl.load_workbook(self.fn)
            self.ws = f.active
            self.spreadsheet_data = self.parse_workbook(f.active)
        self.et = exiftool.ExifTool('C:\\Users\\evharley\\Downloads\\exiftool-11.32\\exiftool.exe')
       
        
    def parse_workbook(self, ws):
        keys = {ws[1][col].value: col for col in range(ws.max_column)}
        spreadsheet_data = {}
        for row in range(2, ws.max_row):
            print('Row: {}'.format(row))
            folder = ws[row][keys['Directory']].value
            filename = ws[row][keys['file_name']].value
            filepath = '{0}/{1}{2}'.format(self.directory,folder, filename)
            workbook_values = {key: ws[row][keys[key]].value for key in keys.keys()}
            keyword_keys = ['Catalogue Number', 'Caption', 'subject','Status',
                            'Conservation Catalogue Sets']
            keywords = ['{0}: {1}'.format(key, workbook_values[key]) 
                        for key in keyword_keys if workbook_values[key] is not None]
            if workbook_values['Conservation Specific'] is not None:
                keywords.extend(workbook_values['Conservation Specific'].strip("[]'").split(','))
            comments = [workbook_values[key] for key in ['rights', 'Comment'] if workbook_values[key] is not None]
            authors = [workbook_values[key] for key in ['creator', 'Author'] if workbook_values[key] is not None]
            spreadsheet_data[filepath] = {'keywords': set(keywords), 'comments': set(comments), 'authors': set(authors)}
        pickle.dump(spreadsheet_data, open(self.tail[:self.tail.find('.')] + '.pkl', 'wb'))
        return spreadsheet_data

    def _write_data(self):
        self.et.start()
        for directory, subdirectories, files in os.walk(self.directory):
            for file in files:
                if not any(file.lower().endswith(ending) for ending in ('.tif', '.jpg', '.nef')):
                    continue
                filepath = os.path.join(directory, file).replace('\\', '/')
                results = self._handle_data(filepath)
                filename = bytes(filepath, 'utf-8')
                if results is None:
                    continue
                for result in results:
                    self.et.execute(results[result], filename)


        self.et.terminate()
        return None

    def _handle_data(self, filepath):
        xmp_data = self._get_xmp_data(filepath)
        
        if filepath in self.spreadsheet_data.keys():
            keywords = self.spreadsheet_data[filepath]['keywords']
            comments = self.spreadsheet_data[filepath]['comments']
            authors = self.spreadsheet_data[filepath]['authors']
        else:
            keywords = []
            comments = []
            authors = []

        if xmp_data == {} and filepath not in self.spreadsheet_data.keys():
            return None
        xmp_data_paths = {'XMP:Method': 'keywords', 
                          'XMP:License': 'keywords', 
                          'XMP:Topicresponsibility': 'comments', 
                          'XMP:Date_entered': 'keywords', 
                          'XMP:Label': 'keywords', 
                          'XMP:History': 'keywords', 
                          'XMP:Status': 'keywords', 
                          'XMP:TopicHumanHistory': 'comments', 
                          'XMP:Description': 'keywords', 
                          'XMP:SubjectArea': 'keywords', 
                          'XMP:BordenNumber': 'keywords', 
                          'XMP:Author': 'authors', 
                          'XMP:TopicFacilities': 'comments', 
                          'XMP:UsageTerms': 'comments', 
                          'XMP:CallNumber': 'keywords', 
                          'XMP:Writer': 'authors', 
                          'XMP:Subject': 'keywords', 
                          'XMP:Comment': 'comments', 
                          'XMP:Artist': 'authors', 
                          'XMP:Keywords': 'keywords', 
                          'XMP:Caption': 'keywords', 
                          'XMP:Creator': 'authors', 
                          'XMP:ConservationCatalogSets': 'keywords', 
                          'XMP:CatalogueNumber': 'keywords', 
                          'XMP:UserComment': 'comments', 
                          'XMP:TopicNaturalHistory': 'comments', 
                          'XMP:ModifyDate': 'keywords', 
                          'XMP:Event': 'keywords', 
                          'XMP:Title': 'keywords', 
                          'XMP:OwnerName': 'keywords',
                          'XMP:AccessionNumber': 'keywords', 
                          'XMP:Rights': 'comments', 
                          'XMP:State': 'comments', 
                          'XMP:Copyright': 'comments'}

        for datum in xmp_data.keys():
            if xmp_data_paths[datum] == 'keywords':
                list_to_compare = list(keywords)
            elif xmp_data_paths[datum] == 'comments':
                list_to_compare = list(comments)
            else:
                list_to_compare = list(authors)

            if not isinstance(xmp_data[datum], list):
               value = xmp_data[datum]
            else:
                value = ', '.join([str(item) for item in xmp_data[datum]])
            if any(str(value) in item for item in list_to_compare) or \
                any(datum[4:] in item for item in list_to_compare):
                continue
            else:
                 list_to_compare.append('{0}: {1}'.format(datum[4:], value))
            if xmp_data_paths[datum] == 'keywords':
                keywords = set(list_to_compare)
            elif xmp_data_paths[datum] == 'comments':
                comments = set(list_to_compare)
            else:
                authors = set(list_to_compare)
            
        return_data = {'keywords': bytes('{0}={1}'.format('-EXIF:XPKeywords', '; '.join(keywords)), 'utf-8'),
                       'comments': bytes('{0}={1}'.format('-EXIF:XPComment', '; '.join(comments)), 'utf-8'),
                       'authors': bytes('{0}={1}'.format('-EXIF:XPAuthor', '; '.join(authors)), 'utf-8')}
        return return_data

    def _get_xmp_data(self, filepath):
        try:
            metadata = self.et.get_metadata(filepath)
        except:
            return {}
        metadata_to_keep =  {'XMP:Method', 'XMP:License', 'XMP:Topicresponsibility', 'XMP:Date_entered', 
                             'XMP:Label', 'XMP:History', 'XMP:Status', 'XMP:TopicHumanHistory', 
                             'XMP:Description', 'XMP:SubjectArea', 'XMP:BordenNumber', 'XMP:Author', 
                             'XMP:TopicFacilities', 'XMP:UsageTerms', 'XMP:CallNumber', 'XMP:Writer', 
                             'XMP:Subject', 'XMP:Comment', 'XMP:Artist', 'XMP:Keywords', 'XMP:Caption', 
                             'XMP:Creator', 'XMP:ConservationCatalogSets', 'XMP:CatalogueNumber', 
                             'XMP:UserComment', 'XMP:TopicNaturalHistory', 'XMP:ModifyDate', 'XMP:Event', 
                             'XMP:Title', 'XMP:OwnerName', 'XMP:AccessionNumber', 'XMP:Rights', 'XMP:State', 
                             'XMP:Copyright'}
        xmp_data = {key: metadata[key] for key in metadata.keys() if key in metadata_to_keep}

        return xmp_data

    def gather_xmp_stats(self):
        self.et.start()
        keys = {self.ws[1][col].value: col for col in range(self.ws.max_column)}
        list_of_fields_used = set()
        for row in range(2, self.ws.max_row):
            folder = self.ws[row][keys['Directory']].value
            folder = folder[folder.find('/') + 1:]
            filename = self.ws[row][keys['file_name']].value
            filepath = '{0}/{1}{2}'.format(self.directory,folder, filename)
            try:
                xmp_data = self._get_xmp_data(filepath)
            except:
                continue
            fields_used = {key for key in xmp_data.keys() if xmp_data[key] is not None}
            list_of_fields_used = list_of_fields_used.union(fields_used)
        self.et.terminate()
        print(list_of_fields_used)
        return list_of_fields_used

    def delete_duplicates(self):
        for directory, subdirectories, files in os.walk(self.directory):
            for file in files:
                if file.endswith('_original'):
                    os.remove(os.path.join(directory, file))
                    print('{0} removed'.format(os.path.join(directory, file)))
        return 0


if __name__ == '__main__':
    writer = MetadataWriter()
    


    
   